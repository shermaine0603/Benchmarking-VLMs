import ollama
import numpy as np
import pandas as pd
import openpyxl
import time
from sklearn.metrics import confusion_matrix

def get_model(model):

    model_versions = {
        'llava': 'llava:7b',
        'qwen': 'qwen2.5vl:3b',
        'gemma': 'gemma3:4b'
    }
    return model_versions[model]

valid_models = ['llava', 'qwen', 'gemma']
while True:
    model = input("Please pick the model you want to use (llava or qwen or gemma): ").strip().lower()
    if model in valid_models:
        break
    else:
        print("Invalid choice. Please type exactly: llava, qwen, or gemma. \n")

model_name = get_model(model)

excel_file = "excel_file.xlsx"
folder_name = "images_folder"
img_dataset_direct = "./" + folder_name + "/"

excel_data = pd.read_excel(excel_file)
wb = openpyxl.load_workbook(excel_file)
ws = wb.active
start_time = time.time()

for i in range(len(excel_data)):
    image_name = str(excel_data['Image'][i])+".jpg"
    query = str(excel_data['Query'][i])

    res = ollama.chat(
        model=model_name,
        messages=[
            {
                'role': 'user',
                'content': query+"? Please answer with only Yes or No, with no explanations.",
                'images': [img_dataset_direct + image_name]
            }
        ]
    )

    reply = (res['message']['content'])
    cell_no = "E" + str(i + 2)
    if "yes" in reply.lower():
        ws[cell_no] = 'yes'
    elif "no" in reply.lower():
        ws[cell_no] = 'no'
    else:
        ws[cell_no] = ''

    wb.save(excel_file)
    print(i+1)
    print(reply)
    print("\n")

print("It works!")
print("Model:", model_name)
print("Dataset:", folder_name)
end_time = time.time()
total_time = end_time - start_time
print(f"\nTotal time taken: {total_time:.2f} seconds")

excel_data = pd.read_excel(excel_file)
print(len(excel_data))

cm = confusion_matrix(excel_data['Answer'], excel_data['Generated Reply'], labels=['no', 'yes'])
TN, FP, FN, TP = cm[0,0], cm[0,1], cm[1,0], cm[1,1]

# Create readable table
df = pd.DataFrame(
    [[TN, FP], [FN, TP]],
    index=["Actual No", "Actual Yes"],
    columns=["Predicted No", "Predicted Yes"]
)

print("Confusion Matrix:")
print(df)
print("\n")

Accuracy = (TN+TP)/(TN+FP+FN+TP)
print("Model's Classification Accuracy :", round(Accuracy,3))

Precision = (TP/(TP+FP))
Recall = (TP/(TP+FN))

print("Model's Precision :", round(Precision,3))
print("Model's Recall :", round(Recall,3))

F1 = (2*Precision*Recall)/(Precision+Recall)
print("F1 Score: ", round(F1,3))

Specificity = (TN/(TN+FP))
print("Model's Specificity :", round(Specificity,3))